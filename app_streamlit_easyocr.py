
import streamlit as st
import easyocr
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import tempfile
import os

# Configuration de la page
st.set_page_config(page_title="Extracteur de Factures PDF", page_icon="📄", layout="wide")

# Logo et titre
st.image("https://streamlit.io/images/brand/streamlit-mark-color.png", width=100)
st.title("Extracteur de Factures PDF")

# Option multilingue
lang = st.selectbox("Choisissez la langue / Choose language", ["Français", "English"])

# Upload PDF
uploaded_file = st.file_uploader("Téléversez votre facture PDF / Upload your invoice PDF", type=["pdf"])

if uploaded_file:
    try:
        # Sauvegarde temporaire
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        temp_pdf.write(uploaded_file.read())
        temp_pdf.close()

        # OCR multi-pages
        reader = easyocr.Reader(['fr', 'en'])
        doc = fitz.open(temp_pdf.name)
        result = []
        for page in doc:
            pix = page.get_pixmap()
            img_path = tempfile.NamedTemporaryFile(delete=False, suffix=".png").name
            pix.save(img_path)
            result += reader.readtext(img_path, detail=0)

        if not result:
            st.error("Aucun texte détecté / No text detected")
        else:
            st.subheader("Texte extrait / Extracted text")
            st.text("
".join(result))

            # Export Excel enrichi
            wb = Workbook()
            ws = wb.active
            ws.title = "Extraction"
            ws.append(["Données extraites / Extracted Data"])
            ws["A1"].font = Font(bold=True, color="FFFFFF")
            ws["A1"].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for line in result:
                ws.append([line])

            excel_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
            wb.save(excel_path)

            with open(excel_path, "rb") as f:
                st.download_button("Télécharger en Excel / Download Excel", f, file_name="extraction.xlsx")

            # Bouton pour télécharger le PDF original
            with open(temp_pdf.name, "rb") as f:
                st.download_button("Télécharger le PDF original / Download original PDF", f, file_name="facture_originale.pdf")

    except Exception as e:
        st.error(f"Erreur lors du traitement : {e}")
